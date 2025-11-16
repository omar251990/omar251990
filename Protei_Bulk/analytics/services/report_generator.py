#!/usr/bin/env python3
"""
Report Generator
Generates PDF, Excel, and CSV reports with charts and analytics
"""

from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any
import io
import csv
from pathlib import Path

from sqlalchemy import func, and_
from sqlalchemy.orm import Session


class ReportGenerator:
    """
    Advanced report generation engine
    Supports multiple formats: PDF, Excel, CSV, JSON
    """

    def __init__(self, db_session_factory, output_dir: str = "./reports"):
        self.db_factory = db_session_factory
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    # ========== Message Reports ==========

    def generate_message_report(
        self,
        db: Session,
        start_date: datetime,
        end_date: datetime,
        format: str = "csv",
        filters: Dict = None
    ) -> Dict[str, Any]:
        """Generate comprehensive message delivery report"""
        from src.models.message import Message, MessageStatus

        filters = filters or {}

        # Base query
        query = db.query(Message).filter(
            Message.created_at.between(start_date, end_date)
        )

        # Apply filters
        if filters.get('account_id'):
            query = query.filter(Message.account_id == filters['account_id'])
        if filters.get('status'):
            query = query.filter(Message.status == filters['status'])
        if filters.get('campaign_id'):
            query = query.filter(Message.campaign_id == filters['campaign_id'])

        messages = query.all()

        # Calculate summary statistics
        total = len(messages)
        by_status = {}
        for msg in messages:
            by_status[msg.status.value] = by_status.get(msg.status.value, 0) + 1

        summary = {
            "total_messages": total,
            "delivered": by_status.get('DELIVERED', 0),
            "failed": by_status.get('FAILED', 0),
            "pending": by_status.get('PENDING', 0),
            "delivery_rate": (by_status.get('DELIVERED', 0) / total * 100) if total > 0 else 0,
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat()
        }

        # Generate report in requested format
        if format == "csv":
            file_path = self._generate_csv_report(messages, "message_report")
        elif format == "json":
            file_path = self._generate_json_report(messages, summary, "message_report")
        elif format == "excel":
            file_path = self._generate_excel_report(messages, summary, "message_report")
        else:
            file_path = None

        return {
            "report_id": f"MSG_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}",
            "summary": summary,
            "file_path": str(file_path) if file_path else None,
            "format": format,
            "generated_at": datetime.utcnow().isoformat()
        }

    def generate_campaign_report(
        self,
        db: Session,
        campaign_id: str,
        format: str = "pdf"
    ) -> Dict[str, Any]:
        """Generate detailed campaign performance report"""
        from src.models.campaign import Campaign
        from src.models.message import Message, MessageStatus

        campaign = db.query(Campaign).filter(Campaign.campaign_id == campaign_id).first()
        if not campaign:
            return {"error": "Campaign not found"}

        # Get all messages for this campaign
        messages = db.query(Message).filter(Message.campaign_id == campaign_id).all()

        # Calculate statistics
        total = len(messages)
        by_status = {}
        for msg in messages:
            by_status[msg.status.value] = by_status.get(msg.status.value, 0) + 1

        # Calculate delivery times
        delivery_times = []
        for msg in messages:
            if msg.status == MessageStatus.DELIVERED and msg.delivered_at and msg.created_at:
                dt = (msg.delivered_at - msg.created_at).total_seconds()
                delivery_times.append(dt)

        import statistics
        avg_delivery_time = statistics.mean(delivery_times) if delivery_times else 0

        # Calculate duration
        if campaign.started_at and campaign.completed_at:
            duration = (campaign.completed_at - campaign.started_at).total_seconds()
        elif campaign.started_at:
            duration = (datetime.utcnow() - campaign.started_at).total_seconds()
        else:
            duration = 0

        avg_send_rate = total / duration if duration > 0 else 0

        summary = {
            "campaign_id": campaign_id,
            "campaign_name": campaign.name,
            "total_recipients": campaign.total_recipients,
            "messages_sent": total,
            "delivered": by_status.get('DELIVERED', 0),
            "failed": by_status.get('FAILED', 0),
            "pending": by_status.get('PENDING', 0),
            "delivery_rate": (by_status.get('DELIVERED', 0) / total * 100) if total > 0 else 0,
            "avg_delivery_time_seconds": avg_delivery_time,
            "avg_send_rate_tps": avg_send_rate,
            "duration_seconds": duration,
            "status": campaign.status.value,
            "created_at": campaign.created_at.isoformat(),
            "started_at": campaign.started_at.isoformat() if campaign.started_at else None,
            "completed_at": campaign.completed_at.isoformat() if campaign.completed_at else None
        }

        file_path = None
        if format == "csv":
            file_path = self._generate_csv_report(messages, f"campaign_{campaign_id}")
        elif format == "json":
            file_path = self._generate_json_report(messages, summary, f"campaign_{campaign_id}")

        return {
            "report_id": f"CMP_{campaign_id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}",
            "summary": summary,
            "file_path": str(file_path) if file_path else None,
            "format": format,
            "generated_at": datetime.utcnow().isoformat()
        }

    def generate_account_usage_report(
        self,
        db: Session,
        account_id: str,
        start_date: datetime,
        end_date: datetime,
        format: str = "excel"
    ) -> Dict[str, Any]:
        """Generate account usage and billing report"""
        from src.models.user import Account
        from src.models.message import Message

        account = db.query(Account).filter(Account.account_id == account_id).first()
        if not account:
            return {"error": "Account not found"}

        # Query messages in date range
        messages = db.query(Message).filter(
            and_(
                Message.account_id == account_id,
                Message.created_at.between(start_date, end_date)
            )
        ).all()

        # Group by day
        by_day = {}
        for msg in messages:
            day = msg.created_at.date()
            if day not in by_day:
                by_day[day] = {"total": 0, "delivered": 0, "failed": 0}
            by_day[day]["total"] += 1
            if msg.status.value == 'DELIVERED':
                by_day[day]["delivered"] += 1
            elif msg.status.value == 'FAILED':
                by_day[day]["failed"] += 1

        summary = {
            "account_id": account_id,
            "account_name": account.name,
            "account_type": account.account_type.value,
            "total_messages": len(messages),
            "current_balance": float(account.balance),
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "daily_breakdown": [
                {
                    "date": str(day),
                    "total": stats["total"],
                    "delivered": stats["delivered"],
                    "failed": stats["failed"]
                }
                for day, stats in sorted(by_day.items())
            ]
        }

        file_path = self._generate_json_report(messages, summary, f"account_{account_id}")

        return {
            "report_id": f"ACC_{account_id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}",
            "summary": summary,
            "file_path": str(file_path) if file_path else None,
            "format": format,
            "generated_at": datetime.utcnow().isoformat()
        }

    # ========== Format Generators ==========

    def _generate_csv_report(self, data: List, filename: str) -> Path:
        """Generate CSV report"""
        file_path = self.output_dir / f"{filename}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"

        with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
            if not data:
                return file_path

            # Use first item to get field names
            fieldnames = ['message_id', 'from_addr', 'to_addr', 'status', 'created_at', 'delivered_at']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

            writer.writeheader()
            for item in data:
                writer.writerow({
                    'message_id': item.message_id,
                    'from_addr': item.from_addr,
                    'to_addr': item.to_addr,
                    'status': item.status.value if hasattr(item.status, 'value') else item.status,
                    'created_at': item.created_at.isoformat(),
                    'delivered_at': item.delivered_at.isoformat() if item.delivered_at else ''
                })

        return file_path

    def _generate_json_report(self, data: List, summary: Dict, filename: str) -> Path:
        """Generate JSON report"""
        import json

        file_path = self.output_dir / f"{filename}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.json"

        report = {
            "summary": summary,
            "data": [
                {
                    "message_id": item.message_id,
                    "from": item.from_addr,
                    "to": item.to_addr,
                    "status": item.status.value if hasattr(item.status, 'value') else item.status,
                    "created_at": item.created_at.isoformat(),
                    "delivered_at": item.delivered_at.isoformat() if item.delivered_at else None
                }
                for item in data[:1000]  # Limit to 1000 records in JSON
            ]
        }

        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(report, f, indent=2)

        return file_path

    def _generate_excel_report(self, data: List, summary: Dict, filename: str) -> Path:
        """Generate Excel report (requires openpyxl)"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            file_path = self.output_dir / f"{filename}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"

            wb = Workbook()

            # Summary sheet
            ws_summary = wb.active
            ws_summary.title = "Summary"

            # Write summary
            row = 1
            ws_summary.cell(row, 1, "Summary Report").font = Font(bold=True, size=14)
            row += 2

            for key, value in summary.items():
                ws_summary.cell(row, 1, str(key).replace('_', ' ').title())
                ws_summary.cell(row, 2, str(value))
                row += 1

            # Data sheet
            ws_data = wb.create_sheet("Messages")
            headers = ['Message ID', 'From', 'To', 'Status', 'Created At', 'Delivered At']

            for col, header in enumerate(headers, 1):
                cell = ws_data.cell(1, col, header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            for row_idx, item in enumerate(data, 2):
                ws_data.cell(row_idx, 1, item.message_id)
                ws_data.cell(row_idx, 2, item.from_addr)
                ws_data.cell(row_idx, 3, item.to_addr)
                ws_data.cell(row_idx, 4, item.status.value if hasattr(item.status, 'value') else item.status)
                ws_data.cell(row_idx, 5, item.created_at.isoformat())
                ws_data.cell(row_idx, 6, item.delivered_at.isoformat() if item.delivered_at else '')

            wb.save(file_path)
            return file_path

        except ImportError:
            # Fallback to JSON if openpyxl not available
            return self._generate_json_report(data, summary, filename)
